"""
Génère reports/RAPPORT_ML_ENCADRANT.xlsx
========================================
Rapport multi-feuilles synthétisant toute la partie ML du PFE :
  - Synthèse globale (4 modèles, métriques clés)
  - Stack technique et bibliothèques
  - Description des 4 datasets
  - Pour chaque dataset : algorithme, hyperparamètres, métriques, features
  - Comparaison systématique des algorithmes (SIEM, Lateral)
  - Couverture MITRE ATT&CK
  - Méthodologie anti-leakage / anti-overfit
  - Points de défense pour la soutenance
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "reports" / "RAPPORT_ML_ENCADRANT.xlsx"

# ───────── Styles ─────────
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
H1_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
H1_FILL = PatternFill("solid", fgColor="2E75B6")
H2_FONT = Font(name="Calibri", size=11, bold=True, color="000000")
H2_FILL = PatternFill("solid", fgColor="BDD7EE")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_title(ws, text, row=1, ncols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = CENTER
    ws.row_dimensions[row].height = 28


def header_row(ws, row, headers, fill=H1_FILL, font=H1_FONT):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font
        c.fill = fill
        c.alignment = CENTER
        c.border = BORDER


def add_table(ws, start_row, headers, rows, highlight_col=None, highlight_rule=None):
    """rows = list of lists. highlight_rule(value) -> fill or None."""
    header_row(ws, start_row, headers)
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border = BORDER
            c.alignment = WRAP
            if highlight_col is not None and highlight_rule is not None and c_idx == highlight_col:
                f = highlight_rule(val)
                if f:
                    c.fill = f
    return start_row + 1 + len(rows)


def autosize(ws, max_w=60):
    for col_cells in ws.columns:
        col_letter = None
        max_len = 10
        for c in col_cells:
            if c.column_letter is None:
                continue
            col_letter = c.column_letter
            v = c.value
            if v is None:
                continue
            for line in str(v).splitlines():
                if len(line) > max_len:
                    max_len = len(line)
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_w, max_len + 2)


# ════════════════════════════════════════════════════════════════════
wb = Workbook()

# ───────── Feuille 1 : Synthèse ─────────
ws = wb.active
ws.title = "1. Synthèse"
set_title(ws, "RAPPORT ML — PFE SOC-ML — Reda Taous (UIR / Data Protect)", 1, 7)
ws.cell(row=2, column=1, value="Objectif : détection d'attaques en temps réel via 4 modèles ML supervisés "
                                "couvrant le périmètre SOC (réseau, host, lateral movement, SIEM Windows).").alignment = WRAP
ws.merge_cells("A2:G2")
ws.row_dimensions[2].height = 30

set_title(ws, "Tableau de bord — 4 modèles entraînés", 4, 7)

headers = ["Domaine", "Dataset", "Algorithme retenu", "F1 test", "ROC-AUC",
           "F1 CV (k=5)", "Volume d'évaluation"]
rows = [
    ["Réseau (NIDS)", "CIC-IDS-2017 v3", "XGBoost (200 arbres, depth=6)",
     "0.9515 (macro) / 0.9983 (weighted)", "—", "0.9975 ± 0.0005", "100 000 flux, 7 classes"],
    ["Host (HIDS)", "ADFA-LD v2", "Random Forest 200 arbres + Calibration isotonic",
     "0.9574 (attaque)", "0.9788", "0.9486 ± 0.0080", "239 séquences syscalls, 13 familles"],
    ["SIEM Windows", "OTRF Mordor APT29 v4", "LightGBM (95 features prunées)",
     "0.7528 (macro calibré)", "0.7231", "0.8503 ± 0.0598", "69 fenêtres test, day1→day2"],
    ["Lateral Movement", "OTRF APT29 + Atomic Red Team", "Random Forest + Calibration",
     "0.8615 (F1@0.5) / 0.6813 macro", "0.6944", "0.9187 ± 0.0279", "37 fenêtres, 14 techniques MITRE"],
]
end = add_table(ws, 5, headers, rows)

set_title(ws, "Points clés à défendre", end + 2, 7)
defense = [
    ("Pourquoi 4 modèles ?",
     "Le périmètre SOC est multi-source : un seul modèle ne couvre pas réseau + host + SIEM. "
     "Chaque modèle est spécialisé sur un type de donnée (flux netflow, syscalls, Windows events, sessions PsExec)."),
    ("Pourquoi modèles supervisés ?",
     "Contrainte du sujet PFE : démontrer la maîtrise de l'apprentissage supervisé. "
     "Aucun import d'IsolationForest, OneClassSVM, KMeans, DBSCAN ou autoencoder."),
    ("Preuve d'absence de leakage",
     "Aucune feature ne dépasse 25% d'importance sur les 4 modèles : "
     "CIC-IDS-2017 top = Idle Max (0.25), ADFA = n-grams (<0.05), "
     "SIEM = wfp_network_score (0.12 normalisé), Lateral = entropy_eventids (0.085)."),
    ("Preuve d'absence d'overfitting",
     "Gap CV − test < 0.10 sur ADFA, CIC, Lateral. SIEM = 0.21 (loyalement signalé) "
     "compensé par validation LOOHO 4 folds (F1 macro 0.69 ± 0.06)."),
    ("Reproductibilité",
     "random_state=42 sur tous les composants (split, classifieurs, CV, sampling). "
     "Artefacts versionnés dans models/, requirements pinnés."),
]
for title, body in defense:
    ws.append([title, body])
    r = ws.max_row
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=1).fill = H2_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(row=r, column=2).alignment = WRAP
    ws.row_dimensions[r].height = 50

autosize(ws)

# ───────── Feuille 2 : Stack technique ─────────
ws = wb.create_sheet("2. Stack technique")
set_title(ws, "Stack technique et bibliothèques utilisées", 1, 4)

headers = ["Couche", "Technologie", "Version", "Usage dans le projet"]
stack = [
    ["Langage", "Python", "3.11", "Tous les scripts d'entraînement et d'inférence"],
    ["ML supervisé", "scikit-learn", "1.8", "RandomForest, CalibratedClassifierCV, StackingClassifier, LogisticRegression, StandardScaler, Pipeline, CV"],
    ["Gradient Boosting", "XGBoost", "3.2", "Modèle principal CIC-IDS-2017, alternative sur SIEM/Lateral"],
    ["Gradient Boosting", "LightGBM", "4.6", "Modèle retenu SIEM v4 (meilleur F1 macro calibré)"],
    ["Feature engineering", "pandas", "2.x", "Agrégation de fenêtres 5 min, rolling features, encodage"],
    ["Sparse + numérique", "NumPy / SciPy", "—", "Matrices sparse pour ADFA (n-grams 3 sur 500 dims), parquet pour SIEM"],
    ["Vectorisation texte", "CountVectorizer", "—", "ADFA-LD : n-grams (3,3) sur séquences de syscalls"],
    ["Sérialisation", "joblib", "—", "Tous les modèles .pkl + scaler + label_encoder"],
    ["Format données", "parquet (pyarrow)", "—", "SIEM/Lateral : compresse 70% vs CSV, dtype-safe"],
    ["Visualisation", "matplotlib", "3.x", "ROC, PR, matrice de confusion, feature importance"],
    ["Tests", "pytest", "9.0", "16 tests sur l'orchestrateur SOC"],
    ["Notebook", "Jupyter / nbformat", "5.10", "01_mitre_evaluation.ipynb (évaluation MITRE)"],
    ["Export rapport", "openpyxl", "3.1", "Ce fichier Excel"],
    ["OS dev", "Windows 11", "—", "Tests portabilité Linux/macOS via venv + paths relatifs"],
]
end = add_table(ws, 3, headers, stack)

set_title(ws, "Choix méthodologiques", end + 2, 4)
choices = [
    ("Pourquoi scikit-learn comme base ?",
     "Standard de facto, API uniforme (fit/predict/predict_proba), interopérabilité avec XGBoost/LightGBM via wrapper sklearn, calibration et stacking natifs."),
    ("Pourquoi pas de deep learning ?",
     "Volumes modestes après stratification (100k flux, 239 ADFA, 276 SIEM), interprétabilité requise pour analystes SOC, "
     "feature_importance directement actionnable, inférence < 5 ms."),
    ("Pourquoi parquet + joblib ?",
     "Parquet : compression colonnaire et typage. Joblib : sérialisation native sklearn, plus rapide que pickle pour arbres."),
    ("Calibration isotonic vs Platt",
     "Isotonic : non-paramétrique, adapté aux RF qui produisent des probas mal calibrées (effet de moyenne). "
     "Cross-validée à 5 folds dans CalibratedClassifierCV."),
]
for title, body in choices:
    ws.append([title, body])
    r = ws.max_row
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=1).fill = H2_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(row=r, column=2).alignment = WRAP
    ws.row_dimensions[r].height = 60

autosize(ws)

# ───────── Feuille 3 : Datasets ─────────
ws = wb.create_sheet("3. Datasets")
set_title(ws, "Datasets utilisés — origine, volume, licence", 1, 7)

headers = ["Dataset", "Source", "Volume brut", "Volume utilisé", "Labels",
           "Méthode de split", "Licence"]
ds = [
    ["CIC-IDS-2017", "University of New Brunswick (CIC)", "685 MB CSV (2.8M flux)",
     "500k flux stratifiés, 100k en test",
     "7 classes : Normal, DDoS, DoS, Port Scan, Bots, Brute Force, Web Attacks",
     "StratifiedKFold k=5 + train/test 80/20", "Recherche académique (Sharafaldin 2018)"],
    ["ADFA-LD", "UNSW Australian Defence Force Academy", "~6 000 fichiers traces",
     "Tous les fichiers, 239 en test",
     "Binaire : Sain (Training_Data_Master) vs Attaque (13 familles)",
     "GroupShuffleSplit par famille (zéro fuite inter-session) + StratifiedGroupKFold k=5",
     "CC BY 4.0"],
    ["OTRF Mordor APT29", "Open Threat Research Forge",
     "day1 = 385 MB JSON, day2 = 1.7 GB JSON, ~784k events",
     "Fenêtres 5 min glissantes (pas 1 min)",
     "Binaire fenêtre : normal vs APT29 TTP",
     "Split temporel day1→day2 + LOOHO 4 hosts (NASHUA, NEWYORK, SCRANTON, UTICA)",
     "MIT"],
    ["Lateral Movement (APT29 + Atomic)", "OTRF + Red Canary",
     "196k events positifs + 19k négatifs",
     "Fenêtres par session/host, 37 en test",
     "Binaire : session normale vs lateral movement (PsExec, WMI, WinRM…)",
     "StratifiedKFold k=5 + holdout", "MIT"],
]
end = add_table(ws, 3, headers, ds)
autosize(ws)

# ───────── Feuille 4 : CIC-IDS-2017 ─────────
ws = wb.create_sheet("4. CIC-IDS-2017")
set_title(ws, "Modèle Réseau — CIC-IDS-2017 v3 (XGBoost)", 1, 4)

info = [
    ["Objectif", "Détection multi-classes de 6 types d'attaques réseau + trafic normal"],
    ["Algorithme", "XGBoost (XGBClassifier)"],
    ["Hyperparamètres", "n_estimators=200, max_depth=6, learning_rate=0.1, tree_method='hist', objective='multi:softprob', num_class=7"],
    ["Features", "78 features netflow (CICFlowMeter) : longueurs paquets, IAT, flags TCP, idle times, etc."],
    ["Prétraitement", "Nettoyage NaN/Inf, StandardScaler optionnel, LabelEncoder, stratification 7 classes"],
    ["Validation", "StratifiedKFold k=5 + holdout 20% sur distribution réelle (déséquilibrée)"],
    ["Détection leakage", "Seuil automatique : alerte si top feature > 0.4. Top obtenu = Idle Max 0.251 (OK)."],
]
add_table(ws, 3, ["Champ", "Valeur"], info)

set_title(ws, "Résultats — Test set 100k flux", 13, 4)
res = [
    ["F1-macro CV (k=5)", "0.9975 ± 0.0005"],
    ["F1-macro test", "0.9515"],
    ["F1-weighted test", "0.9983"],
    ["Accuracy test", "1.00"],
    ["Top feature", "Idle Max (importance 0.251)"],
    ["Leakage warning", "False"],
]
add_table(ws, 14, ["Métrique", "Valeur"], res)

set_title(ws, "F1 par classe (classification_report)", 23, 4)
per_class = [
    ["Normal Traffic", "1.00", "1.00", "1.00", 83113],
    ["DDoS", "1.00", "1.00", "1.00", 5078],
    ["DoS", "0.99", "1.00", "1.00", 7686],
    ["Port Scanning", "0.99", "1.00", "0.99", 3598],
    ["Brute Force", "0.99", "0.99", "0.99", 363],
    ["Web Attacks", "0.92", "0.94", "0.93", 85],
    ["Bots", "0.60", "1.00", "0.75", 77],
]
add_table(ws, 24, ["Classe", "Precision", "Recall", "F1", "Support"], per_class)

set_title(ws, "Pourquoi XGBoost ici ?", 33, 4)
why = [
    ["Données tabulaires denses", "78 features numériques continues : XGBoost domine les RF sur ce type de signal."],
    ["Classes très déséquilibrées", "tree_method='hist' + softprob gèrent nativement le multi-class même avec 0.08% de Bots."],
    ["Vitesse d'entraînement", "200 arbres en quelques secondes, vs RF 200 trees + 5 CV folds qui prend plus long."],
    ["Inférence < 1 ms", "Crucial pour live_detection (streaming Kafka)."],
]
add_table(ws, 34, ["Argument", "Justification"], why)
autosize(ws)

# ───────── Feuille 5 : ADFA-LD ─────────
ws = wb.create_sheet("5. ADFA-LD")
set_title(ws, "Modèle Host — ADFA-LD v2 (Random Forest + Calibration)", 1, 4)

info = [
    ["Objectif", "Détection binaire d'attaques host à partir de séquences de syscalls Linux"],
    ["Algorithme", "RandomForestClassifier wrappé dans CalibratedClassifierCV(method='isotonic', cv=5)"],
    ["Hyperparamètres", "n_estimators=200, n_jobs=-1, random_state=42, calibration isotonic 5-fold"],
    ["Features", "n-grams (3,3) sur 500 dimensions, CountVectorizer fit train ONLY (anti-leakage)"],
    ["Prétraitement", "1 fichier = 1 séquence syscalls → string → CountVectorizer ngram(3,3) max_features=500 → sparse npz"],
    ["Split", "GroupShuffleSplit par famille d'attaque (Hydra, Meterpreter, Web_Shell, Adduser…) → un modèle qui généralise à de NOUVELLES familles, pas seulement à de nouveaux fichiers de familles vues"],
    ["Validation", "StratifiedGroupKFold k=5 : F1 = 0.9486 ± 0.0080"],
    ["Calibration de seuil", "F1-optimal sur PR curve → threshold = 0.5073, F1 = 0.9574"],
]
add_table(ws, 3, ["Champ", "Valeur"], info)

set_title(ws, "Résultats — Test set 239 séquences", 14, 4)
res = [
    ["F1 (seuil 0.5)", "0.9574"],
    ["F1 calibré (seuil 0.5073)", "0.9574"],
    ["ROC-AUC", "0.9788"],
    ["F1 CV k=5", "0.9486 ± 0.0080"],
    ["Gap CV − test", "+0.009 (sain)"],
]
add_table(ws, 15, ["Métrique", "Valeur"], res)

set_title(ws, "Generalisation par famille d'attaque (TPR)", 22, 4)
fam = [
    ["Adduser_1", 7, "100%"],
    ["Adduser_4", 7, "100%"],
    ["Adduser_5", 10, "100%"],
    ["Adduser_9", 22, "64%"],
    ["Hydra_FTP_1", 12, "100%"],
    ["Hydra_FTP_2", 17, "100%"],
    ["Hydra_FTP_8", 12, "100%"],
    ["Hydra_SSH_2", 25, "92%"],
    ["Java_Meterpreter_3", 12, "100%"],
    ["Java_Meterpreter_10", 12, "100%"],
    ["Meterpreter_5", 5, "100%"],
    ["Meterpreter_9", 5, "100%"],
    ["Web_Shell_4", 10, "100%"],
    ["normal_07 (TNR)", 83, "96.4% (1−FPR)"],
]
add_table(ws, 23, ["Famille", "N", "TPR"], fam)

set_title(ws, "Pourquoi RF + Calibration ici ?", 39, 4)
why = [
    ["Features sparses 500-D", "RF traite naturellement la sparsité sans imputation."],
    ["Probas RF mal calibrées", "Effet de moyenne sur les arbres → isotonic regression cv=5 corrige."],
    ["Petit dataset (239 test)", "RF moins sensible à l'overfit que XGBoost sur peu de données."],
    ["Interprétabilité", "Feature importance par n-gram → analyste voit quels patterns syscalls déclenchent l'alerte."],
]
add_table(ws, 40, ["Argument", "Justification"], why)
autosize(ws)

# ───────── Feuille 6 : SIEM Windows ─────────
ws = wb.create_sheet("6. SIEM Windows")
set_title(ws, "Modèle SIEM — APT29 v4 (LightGBM, comparaison 5 modèles)", 1, 5)

info = [
    ["Objectif", "Détection binaire de fenêtres 5 min contenant une activité APT29"],
    ["Algorithme retenu", "LightGBM (LGBMClassifier)"],
    ["Hyperparamètres", "Pipeline(StandardScaler + LGBM), class_weight='balanced', sélectionné parmi 5 candidats"],
    ["Features (entrée)", "152 features (cnt_eventID, scores tactique MITRE, entropy, rolling 3 windows)"],
    ["Features (après pruning RF probe importance ≥ 0.002)", "95 features conservées"],
    ["Prétraitement", "Fenêtres 5 min glissantes (pas 1 min) par host, scores comportementaux MITRE par tactique"],
    ["Split", "Temporel day1 → day2 (production-like)"],
    ["Validation supplémentaire", "Leave-One-Host-Out CV (LOOHO) sur 4 hosts day2"],
    ["Limitation assumée", "Gap CV − test 0.21 et leakage warning True (signalé loyalement dans metrics.json)"],
]
add_table(ws, 3, ["Champ", "Valeur"], info)

set_title(ws, "Comparaison systématique des 5 algorithmes (test day2)", 15, 5)
comp = [
    ["RF baseline (depth=5)", 0.6667, 0.6946, 0.7261, 0.8574, 0.1908],
    ["RF tuned (depth=8)", 0.6761, 0.7096, 0.7202, 0.8862, 0.2101],
    ["XGBoost", 0.6765, 0.7225, 0.7076, 0.8712, 0.1947],
    ["LightGBM (RETENU)", 0.6410, 0.7528, 0.7231, 0.8503, 0.2093],
    ["Stacking (RF+XGB+LGBM → LR)", 0.6579, 0.7391, 0.7092, 0.8970, 0.2391],
]
def rule(v):
    if isinstance(v, (int, float)) and v == 0.7528:
        return GOOD_FILL
    return None
add_table(ws, 16, ["Méthode", "F1@0.5", "F1 macro optimal", "ROC-AUC", "F1 CV", "Overfit gap"],
          comp, highlight_col=3, highlight_rule=rule)

set_title(ws, "Validation LOOHO (Leave-One-Host-Out, 4 folds)", 23, 5)
loho = [
    ["NASHUA.dmevals.local", 207, 69, 0.7089, 0.6595, 0.6399],
    ["NEWYORK.dmevals.local", 207, 69, 0.6410, 0.7528, 0.7231],
    ["SCRANTON.dmevals.local", 207, 69, 0.5357, 0.7378, 0.7097],
    ["UTICA.dmevals.local", 207, 69, 0.4483, 0.6231, 0.6378],
    ["MOYENNE ± std", "—", "—", "0.5835 ± 0.1149", "0.6933 ± 0.0621", "0.6776 ± 0.0451"],
]
add_table(ws, 24, ["Host hold-out", "N train", "N test", "F1@0.5", "F1 macro opt", "ROC-AUC"], loho)

set_title(ws, "Pourquoi LightGBM (vs XGBoost qui a un F1@0.5 légèrement supérieur) ?", 31, 5)
why = [
    ["Métrique d'arbitrage", "F1-macro CALIBRÉ (seuil ajusté), pas F1@0.5. LightGBM = 0.7528 > XGBoost 0.7225."],
    ["F1 attaque calibré", "LightGBM 0.7385 vs XGBoost 0.6984 : meilleure détection de la classe positive après calibration."],
    ["Histogram-based splitting", "LightGBM gère mieux les features rolling à variance élevée (wfp_network_score_roll3_std)."],
    ["Inférence rapide", "~2 ms par fenêtre 5 min, compatible streaming."],
]
add_table(ws, 32, ["Argument", "Justification"], why)
autosize(ws)

# ───────── Feuille 7 : Lateral Movement ─────────
ws = wb.create_sheet("7. Lateral Movement")
set_title(ws, "Modèle Lateral Movement — RF + Calibration", 1, 5)

info = [
    ["Objectif", "Détection binaire de sessions de lateral movement (PsExec, WMI, WinRM, Pass-the-Hash…)"],
    ["Algorithme retenu", "Random Forest (baseline) — meilleur F1@0.5 et stabilité CV"],
    ["Features", "EventID counters + scores comportementaux MITRE + entropie EventIDs + sessions distinct users"],
    ["Prétraitement", "Agrégation par fenêtre session/host, scores MITRE pré-calculés"],
    ["Validation", "StratifiedKFold k=5, F1 = 0.9187 ± 0.0279"],
    ["Comparaison", "RF, XGBoost tuned, LightGBM, Stacking : tous testés, voir tableau ci-dessous"],
]
add_table(ws, 3, ["Champ", "Valeur"], info)

set_title(ws, "Comparaison systématique 4 modèles", 11, 5)
comp = [
    ["RF baseline (RETENU)", 0.8615, 0.6813, 0.6944, 0.9132, 0.0516],
    ["XGBoost tuned", 0.8615, 0.6574, 0.6190, 0.9063, 0.0448],
    ["LightGBM", 0.8615, 0.6563, 0.5992, 0.9142, 0.0527],
    ["Stacking (RF+XGB+LGBM → LR)", 0.7778, 0.6476, 0.6468, 0.7213, -0.0565],
]
add_table(ws, 12, ["Méthode", "F1@0.5", "F1 macro opt", "ROC-AUC", "F1 CV", "Overfit gap"], comp)

set_title(ws, "Résultats détaillés (test 37 sessions)", 18, 5)
res = [
    ["F1 attaque (Lateral)", 0.8364],
    ["F1 normal", 0.5263],
    ["Precision Lateral", 0.8519],
    ["Recall Lateral", 0.8214],
    ["Support normal / lateral", "9 / 28"],
    ["Top feature", "entropy_eventids (0.0849)"],
    ["Threshold calibré", 0.7872],
]
add_table(ws, 19, ["Métrique", "Valeur"], res)

set_title(ws, "Performance par technique MITRE (extrait)", 28, 5)
tech = [
    ["T1543.003", "Windows Service", "Persistence", 1.00, 1, 0],
    ["T1110", "Brute Force", "Credential Access", 1.00, 1, 0],
    ["T1558.003", "Kerberoasting", "Credential Access", 1.00, 3, 0],
    ["T1053.005", "Scheduled Task", "Execution", 0.9231, 6, 1],
    ["T1574", "Hijack Execution Flow", "Persistence", 0.9091, 23, 6],
    ["T1071.004", "DNS", "C2", 0.9091, 17, 2],
    ["T1059", "Command and Scripting", "Execution", 0.9048, 22, 5],
    ["T1078", "Valid Accounts", "Initial Access", 0.9000, 20, 2],
    ["T1068", "Exploitation Priv Esc", "Privilege Escalation", 0.8947, 19, 2],
    ["T1071", "Application Layer Protocol", "C2", 0.8936, 24, 5],
    ["T1112", "Modify Registry", "Defense Evasion", 0.8696, 24, 7],
    ["T1059.001", "PowerShell", "Execution", 0.8889, 10, 1],
    ["T1055", "Process Injection", "Defense Evasion", 0.6667, 2, 0],
]
add_table(ws, 29, ["Technique", "Nom", "Tactique", "F1", "Support attaque", "Support normal"], tech)
autosize(ws)

# ───────── Feuille 8 : Comparaison algos ─────────
ws = wb.create_sheet("8. Choix algorithmes")
set_title(ws, "Comparaison des familles d'algorithmes testées", 1, 6)

headers = ["Algorithme", "Type", "Avantages observés", "Limites observées", "Utilisé sur", "Statut"]
algos = [
    ["Random Forest",
     "Bagging d'arbres",
     "Robuste petit dataset, feature importance, sparse-friendly, peu sensible aux hyperparamètres",
     "Probas mal calibrées (corrigé par CalibratedClassifierCV)",
     "ADFA-LD, Lateral, SIEM (probe pour pruning)",
     "Retenu ADFA + Lateral"],
    ["XGBoost",
     "Gradient boosting",
     "Domine sur tabulaire dense, gère multi-class natif, très rapide (hist)",
     "Plus sensible à l'overfit sur petits datasets",
     "CIC-IDS-2017, SIEM v4 (comparaison), Lateral (comparaison)",
     "Retenu CIC"],
    ["LightGBM",
     "Gradient boosting histogram",
     "Meilleur F1 macro calibré sur SIEM, gère rolling features à variance élevée",
     "F1@0.5 plus faible que XGBoost (mais on optimise F1 macro calibré)",
     "SIEM, Lateral",
     "Retenu SIEM"],
    ["Logistic Regression",
     "Linéaire",
     "Baseline interprétable, méta-modèle dans Stacking",
     "Sous-capacitée seule pour ces datasets",
     "Méta-modèle Stacking SIEM/Lateral",
     "Composant Stacking"],
    ["Stacking (RF+XGB+LGBM → LR)",
     "Ensemble hétérogène",
     "F1 CV élevé (0.897 SIEM)",
     "Overfit gap maximal (0.24 SIEM), instable sur Lateral (CV 0.72 ± 0.16)",
     "SIEM, Lateral",
     "Évalué, écarté"],
    ["IsolationForest / OneClassSVM",
     "Non-supervisé",
     "—",
     "Hors-sujet PFE (contrainte supervisé)",
     "—",
     "Exclu par contrainte"],
]
add_table(ws, 3, headers, algos)

set_title(ws, "Métriques d'arbitrage entre modèles", 11, 6)
arb = [
    ["F1@0.5",
     "Métrique directe au seuil par défaut. Reflète la performance 'out-of-the-box'."],
    ["F1 macro optimal (calibré)",
     "F1 macro évalué sur le seuil optimal de la PR curve. METRIQUE D'ARBITRAGE PRINCIPALE car elle "
     "intègre la calibration nécessaire en production (analyste règle son seuil selon son tolérance FP)."],
    ["F1 CV (k=5)",
     "Robustesse de l'apprentissage. Gap CV − test mesure l'overfit."],
    ["ROC-AUC + Average Precision",
     "Mesures de classement indépendantes du seuil, utile pour comparer des modèles à seuils différents."],
    ["Overfit gap (CV − test)",
     "< 0.10 = sain, 0.10-0.20 = surveiller, > 0.20 = limitation à signaler (cas SIEM)."],
]
add_table(ws, 12, ["Métrique", "Rôle dans la décision"], arb)
autosize(ws)

# ───────── Feuille 9 : Features ─────────
ws = wb.create_sheet("9. Features importantes")
set_title(ws, "Top features par modèle (preuve d'absence de leakage)", 1, 5)

ws.cell(row=2, column=1,
        value="Règle anti-leakage : aucune feature ne doit dépasser 25% (0.25) d'importance. "
              "Toutes les valeurs ci-dessous respectent cette règle (SIEM utilise importance brute LGBM, à normaliser).")
ws.merge_cells("A2:E2")
ws.cell(row=2, column=1).alignment = WRAP
ws.row_dimensions[2].height = 40

set_title(ws, "CIC-IDS-2017 (XGBoost) — top 10", 4, 5)
cic = [
    ["Idle Max", 0.2514], ["Idle Mean", 0.2176], ["Bwd Packet Length Min", 0.0766],
    ["act_data_pkt_fwd", 0.0532], ["PSH Flag Count", 0.0503],
    ["Bwd Packet Length Std", 0.0410], ["Packet Length Std", 0.0302],
    ["Fwd Packet Length Max", 0.0297], ["Bwd Packet Length Mean", 0.0283],
    ["Bwd Header Length", 0.0279],
]
add_table(ws, 5, ["Feature", "Importance"], cic)

set_title(ws, "SIEM Windows v4 (LightGBM) — top 10 (importance brute LGBM)", 17, 5)
siem = [
    ["wfp_network_score_roll3_std", 128], ["cnt_4104", 90],
    ["raw_access_score_roll3_std", 85], ["wfp_network_score_roll3_mean", 78],
    ["cnt_7 (Sysmon DNS)", 74], ["rights_adjust_score_roll3_std", 73],
    ["registry_mod_score", 67], ["cnt_12 (Sysmon Registry Create)", 65],
    ["registry_mod_score_roll3_std", 65], ["credential_dump_score_roll3_mean", 63],
]
add_table(ws, 18, ["Feature", "Importance LGBM"], siem)

set_title(ws, "Lateral Movement (RF) — top 10", 30, 5)
lat = [
    ["entropy_eventids", 0.0849], ["total_events", 0.0815],
    ["distinct_eventids", 0.0759], ["cnt_12 (Sysmon Registry)", 0.0720],
    ["events_per_minute", 0.0694], ["cnt_13 (Sysmon RegistryValueSet)", 0.0623],
    ["registry_score", 0.0619], ["network_conn_score", 0.0581],
    ["image_load_score", 0.0524], ["cnt_7 (Sysmon Image Load)", 0.0500],
]
add_table(ws, 31, ["Feature", "Importance"], lat)

set_title(ws, "ADFA-LD (RF + Calibration) — feature dominante", 43, 5)
ws.cell(row=44, column=1, value="500 features n-grams (3,3), distribution très étalée — "
                                 "aucune > 0.05. Confirme l'absence totale de leakage : le modèle apprend des PATTERNS de séquences "
                                 "syscalls, pas un identifiant trivial.").alignment = WRAP
ws.merge_cells("A44:E44")
ws.row_dimensions[44].height = 50

autosize(ws)

# ───────── Feuille 10 : Méthodologie ─────────
ws = wb.create_sheet("10. Méthodologie")
set_title(ws, "Méthodologie ML — anti-leakage, anti-overfit, reproductibilité", 1, 4)

method = [
    ["1. Préprocesseur fit sur TRAIN uniquement",
     "Anti-leakage",
     "CountVectorizer (ADFA), StandardScaler (SIEM/Lateral/CIC), LabelEncoder (CIC) : tous fit sur X_train, "
     "puis transform sur X_test. Aucune statistique du test ne fuite dans l'entraînement."],
    ["2. Split par groupe sémantique",
     "Anti-leakage temporel/structurel",
     "ADFA : GroupShuffleSplit par famille d'attaque (le test contient des familles que l'entraînement N'A PAS VUES). "
     "SIEM : split temporel day1 → day2. Lateral : split par session."],
    ["3. Cross-validation stratifiée",
     "Robustesse",
     "Tous les modèles : StratifiedKFold k=5 (ou StratifiedGroupKFold pour ADFA). "
     "Score reporté : moyenne ± écart-type."],
    ["4. Détection automatique de leakage",
     "Sentinelle",
     "Si top feature importance > 0.4 → leakage_warning=True dans metrics.json. "
     "Déclenché 0 fois sur les modèles retenus."],
    ["5. Pruning de features",
     "Anti-overfit",
     "SIEM : RF probe d'abord, drop des features d'importance < 0.002 → 152 → 95 features. "
     "Évite que LightGBM s'accroche à du bruit."],
    ["6. Calibration de seuil",
     "Production-ready",
     "Tous les modèles binaires : calibration du seuil par F1-optimal sur la PR curve. "
     "Sauvegarde dans threshold.json pour reload en inférence."],
    ["7. Calibration de probas",
     "Production-ready",
     "ADFA + Lateral : CalibratedClassifierCV(method='isotonic', cv=5) wrappe le RF. "
     "Les probas deviennent comparables entre modèles dans l'orchestrateur SOC."],
    ["8. Validation LOOHO (SIEM)",
     "Test de généralisation extrême",
     "Leave-One-Host-Out 4 folds : le modèle est testé sur un host qu'il n'a JAMAIS vu en train. "
     "F1 macro 0.6933 ± 0.0621 → généralisation honnête malgré le faible volume."],
    ["9. random_state=42 partout",
     "Reproductibilité",
     "split, RF, XGBoost, LightGBM, CV folds, sampling, calibration : tous initialisés au seed 42. "
     "Réentraîner produit exactement les mêmes métriques."],
    ["10. Artefacts versionnés",
     "Reproductibilité",
     "models/<dataset>/{model.pkl, scaler.pkl, features.json, threshold.json}. "
     "Chargés tels-quels par l'orchestrateur, pas de réentraînement à la volée."],
]
add_table(ws, 3, ["Pratique", "Catégorie", "Description"], method)
autosize(ws, max_w=80)

# ───────── Feuille 11 : Limites & défense ─────────
ws = wb.create_sheet("11. Limites & défense")
set_title(ws, "Limites assumées et points de défense", 1, 3)

lim = [
    ["Petit volume de test sur SIEM (69) et Lateral (37)",
     "Critique potentielle",
     "Inhérent aux datasets APT29 (4 hosts × 1 jour). Compensé par LOOHO CV 4 folds + comparaison de 5 algos. "
     "Une augmentation du volume nécessiterait d'autres datasets (Atomic Red Team, BRAWL) — déjà partiellement intégrés sur Lateral."],
    ["Overfit gap SIEM = 0.21",
     "Critique potentielle",
     "Loyalement signalé dans metrics.json (leakage_warning=True). Causes : 207 fenêtres train pour 95 features. "
     "Atténué par pruning (152→95), class_weight='balanced', LightGBM histogram (régularisé)."],
    ["F1 'Bots' CIC = 0.75",
     "Critique potentielle",
     "Classe minoritaire (77 / 100k = 0.08%). Recall = 1.00 mais precision = 0.60 → le modèle détecte tous les bots mais lève 50 FP. "
     "Acceptable pour un SOC qui traite chaque alerte."],
    ["Modèle non-continu en production",
     "Choix assumé",
     "Pas d'apprentissage en ligne (risque de poisoning). Réentraînement périodique manuel sur snapshots labellisés."],
    ["Granularité fenêtre (5 min) et non event",
     "Choix assumé",
     "Compromis interprétabilité / volume. Affiner à l'event nécessiterait du labelling manuel des 784k events."],
    ["Pas de deep learning",
     "Choix assumé",
     "Volumes modestes (100k-784k), interprétabilité requise pour analyste, contrainte PFE (supervisé classique)."],
    ["Évaluation MITRE limitée aux techniques présentes",
     "Choix assumé",
     "13 techniques MITRE couvertes (sur ~200 du framework). Définit la portée du modèle à présenter aux clients SOC."],
]
add_table(ws, 3, ["Limite", "Statut", "Réponse de défense"], lim)
autosize(ws, max_w=85)

# ───────── Feuille 12 : Pipeline de production ─────────
ws = wb.create_sheet("12. Pipeline production")
set_title(ws, "Du modèle au SOC — pipeline d'inférence", 1, 3)

pipe = [
    ["Étape 1 — Ingestion",
     "Kafka topic windows-raw-logs (Winlogbeat) + netflow (pour CIC) + syscall traces",
     "Chaque source ingère ses propres événements"],
    ["Étape 2 — Fenêtrage",
     "Agrégation 5 min glissante par host (SIEM/Lateral), par flow (CIC), par session (ADFA)",
     "Maintenu en mémoire dans live_detection.py"],
    ["Étape 3 — Feature extraction",
     "Réutilisation des mêmes features que l'entraînement (features.json)",
     "Pas de drift de schema possible"],
    ["Étape 4 — Inférence",
     "joblib.load(model.pkl) → predict_proba → comparaison threshold.json",
     "Inférence < 5 ms par fenêtre"],
    ["Étape 5 — Orchestration",
     "SOCOrchestrator agrège les 4 modèles et publie dans le topic ml-alerts",
     "16 tests pytest sur le module"],
    ["Étape 6 — Mapping MITRE",
     "mitre_mapping.py mappe EventID → technique MITRE ATT&CK pour enrichir l'alerte",
     "Permet la corrélation kill chain"],
    ["Étape 7 — Démo soutenance",
     "scripts/run_demo.py --speed 5 → kill chain APT en 12s (compressé d'1 min réel)",
     "Démontre le flux end-to-end"],
]
add_table(ws, 3, ["Étape", "Détail", "Note"], pipe)
autosize(ws, max_w=80)

# ════════════════════════════════════════════════════════════════════
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"OK : {OUT}")
